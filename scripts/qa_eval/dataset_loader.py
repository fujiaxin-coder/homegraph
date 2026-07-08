"""Load qa_eval test datasets from JSONL or Excel (.xlsx)."""

from __future__ import annotations

import json
from pathlib import Path

_XLSX_SUFFIXES = {".xlsx", ".xlsm"}

# Excel header (Chinese benchmark export) -> qa_eval JSONL field
_XLSX_COLUMN_MAP = {
    "一级类别": "category_l1",
    "二级类别": "category_l2",
    "难度": "difficulty",
    "语言": "language",
    "规模": "scale",
    "代码仓": "repo",
    "问题类型": "question_type",
    "query": "query",
    "reference_answer": "reference_answer",
    "explore_query": "explore_query",
    "id": "id",
    "level0": "level0",
    "level1": "level1",
}


def load_dataset(path: Path | str) -> list[dict]:
    """Load benchmark rows from JSONL or Excel."""
    p = Path(path).expanduser().resolve()
    if not p.is_file():
        raise FileNotFoundError(f"测试集不存在: {p}")

    suffix = p.suffix.lower()
    if suffix in _XLSX_SUFFIXES:
        return _load_xlsx(p)
    if suffix == ".jsonl" or suffix == ".json":
        return _load_jsonl(p)

    raise ValueError(
        f"不支持的测试集格式: {p.name}（请使用 .jsonl / .json / .xlsx）"
    )


def _load_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _load_xlsx(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError(
            "读取 Excel 测试集需要 openpyxl，请运行: pip install -r scripts/qa_eval/requirements.txt"
        ) from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError(f"Excel 测试集为空: {path}")

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col_map: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        field = _XLSX_COLUMN_MAP.get(h, h if h in _XLSX_COLUMN_MAP.values() else None)
        if field:
            col_map[field] = i

    if "query" not in col_map:
        raise ValueError(
            f"Excel 测试集缺少 query 列（当前表头: {headers}）"
        )

    rows: list[dict] = []
    for row_idx, row_vals in enumerate(
        ws.iter_rows(min_row=2, values_only=True), start=2
    ):
        if not row_vals:
            continue

        def cell(field: str) -> str:
            idx = col_map.get(field)
            if idx is None or idx >= len(row_vals):
                return ""
            val = row_vals[idx]
            if val is None:
                return ""
            return str(val).strip()

        query = cell("query")
        if not query:
            continue

        item: dict = {
            "id": cell("id") or f"D{len(rows) + 1:02d}",
            "query": query,
            "reference_answer": cell("reference_answer"),
        }
        for field in (
            "category_l1",
            "category_l2",
            "difficulty",
            "language",
            "scale",
            "repo",
            "question_type",
            "level0",
            "level1",
        ):
            val = cell(field)
            if val:
                item[field] = val

        rows.append(item)

    wb.close()
    if not rows:
        raise ValueError(f"Excel 测试集没有有效题目: {path}")
    return rows
